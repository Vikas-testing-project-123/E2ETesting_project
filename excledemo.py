import openpyxl
book = openpyxl.load_workbook("C:/Users/DELL/Documents/dataforautomation.xlsx")

sheet = book.active
Dict={}
# cell = sheet.cell(row=1, column=2)
# print(cell.value)
# sheet.cell(row=2, column=2 ).value="Vikash"
#
# print(sheet.cell(row=2, column=2 ).value)
# print(sheet.max_row)
# print(sheet.max_column)
#
# print(sheet["A4"].value)
#
# for i in range(2, sheet.max_row + 1):
#     # if sheet.cell(row=i, column=1).value == test_case_name:
#     for j in range(2, sheet.max_column + 1):  # for coulum count
#         #dict["lastname"]="shetty"
#         Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
#     print(Dict)
class testData():
    def adddata(self):

        for i in range(2, sheet.max_row + 1):
    # if sheet.cell(row=i, column=1).value == "Testcase 2":
            for j in range(2, sheet.max_column + 1):  # for coulum count
            #dict["lastname"]="shetty"
                Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        print(Dict)

obj = testData()
obj.adddata()
print(obj.Dict[2])
