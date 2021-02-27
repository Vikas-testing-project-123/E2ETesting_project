import openpyxl


class homepagedata:
    test_homepage_data = [{"Firstname": "Vikash", "Email": "Email", "Password": "Test12345"},
                          {"Firstname": "Archa", "Email": "arc", "Password": "Test12345"}] # It is by sending the valuse from here

    # data driven from the excel have a problem now
    # @staticmethod
    # def getTestData(test_case_name):
    #     Dict = {}
    #     book = openpyxl.load_workbook("C:/Users/DELL/Documents/dataforautomation.xlsx")
    #     sheet = book.active
    #     for i in range(1, sheet.max_row + 1):
    #         if sheet.cell(row=i, column=1).value == test_case_name:
    #             for j in range(2, sheet.max_column + 1):  # for coulum count
    #                 # dict["lastname"]="shetty"
    #                 Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
    #     return [Dict]